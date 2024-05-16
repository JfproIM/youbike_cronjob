import  json, ssl, urllib.request
import openpyxl
import os
import xlwt
import pandas as pd
import xlsxwriter
import numpy

os.chdir('大二下\統計學一下\期末報告\資料抓取\cron_job')
'''wb = openpyxl.Workbook()
ws = wb.active'''

wb = openpyxl.load_workbook('youbike_cronjob2.xlsx')

url = 'https://tcgbusfs.blob.core.windows.net/dotapp/youbike/v2/youbike_immediate.json'
context = ssl._create_unverified_context()

with urllib.request.urlopen(url, context=context) as jsondata:
    #將JSON進行UTF-8的BOM解碼，並把解碼後的資料載入JSON陣列中
     data = json.loads(jsondata.read().decode('utf-8-sig')) 




name = "臺大"

# youbike_dict = {}
youbike_dict = {'sna': [], 'mday': [],'total': [], 'available_rent_bikes': [], 'available_return_bikes': [], 'updateTime': []}
for i in data:
      if i['sna'].find(name) >= 0:
        # print(i['sna'],'\t',i['mday'],'\t',i['available_rent_bikes'])
        # youbike_dict[i['sna']] = [i['mday'], i['available_rent_bikes']]
        youbike_dict['sna'].append(i['sna'])
        youbike_dict['mday'].append(i['mday'])
        youbike_dict['total'].append(i['total'])
        youbike_dict['available_rent_bikes'].append(str(i['available_rent_bikes']))
        youbike_dict['available_return_bikes'].append(str(i['available_return_bikes']))
        youbike_dict['updateTime'].append(i['updateTime'])


pf = pd.DataFrame(youbike_dict)
writer = pd.ExcelWriter('youbike_cronjob2.xlsx', engine='openpyxl')
# book = openpyxl.load_workbook(writer)
# writer.wb = wb
pf.to_excel(excel_writer=writer, sheet_name='sheet2')
writer.close()


'''writer = pd.ExcelWriter('test_exist.xlsx', engin='openpyxl')
book = load_workbook(writer.path)
writer.book = book
df.to_excel(excel_writer=writer, sheet_name='sheet3')
writer.save()'''

# engine='xlsxwriter'
# engine='openpyxl'

# print(youbike_dict)
print(pd.DataFrame(youbike_dict))

# append_df_to_excel('youbike_cronjob1.xlsx', pf, sheet_name='Sheet1', startrow=None, truncate_sheet=False)

# export_excel(youbike_dict)
# wb.close()
wb.save('youbike_cronjob1.xlsx')
#wb.save('youbike_cronjob1.xlsx')